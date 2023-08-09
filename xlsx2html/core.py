import contextlib
import io
from collections import defaultdict
from typing import List
import json

import openpyxl
import six
import html
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.styles.colors import COLOR_INDEX, aRGB_REGEX
from openpyxl.utils import rows_from_range, column_index_from_string, units
from openpyxl.worksheet.worksheet import Worksheet

from xlsx2html.compat import OPENPYXL_24
from xlsx2html.constants.border import DEFAULT_BORDER_STYLE, BORDER_STYLES
from xlsx2html.format import format_cell
from xlsx2html.utils.image import bytes_to_datauri
from xlsx2html.utils.theme import theme_and_tint_to_rgb

def render_attrs(attrs):
    if not attrs:
        return ""
    return " ".join(
        ['%s="%s"' % a for a in sorted(attrs.items(), key=lambda a: a[0]) if a[1]]
    )


def render_inline_styles(styles):
    if not styles:
        return ""
    return ";".join(
        [
            "%s: %s" % a
            for a in sorted(styles.items(), key=lambda a: a[0])
            if a[1] is not None
        ]
    )


def normalize_color(wb, color):
    if color.type == 'auto':
         return "#000000"
    if color.type == 'theme':
       return "#" + theme_and_tint_to_rgb(wb, color.theme, color.tint)
    
    rgb = None
    if color.type == "rgb":
        rgb = color.rgb
    if color.type == "indexed":
        try:
            rgb = COLOR_INDEX[color.indexed]
        except IndexError:
            # The indices 64 and 65 are reserved for the system
            # foreground and background colours respectively
            pass
        if not rgb or not aRGB_REGEX.match(rgb):
            # TODO system fg or bg
            rgb = "00000000"
    if rgb:
        return "#" + rgb[2:]
    return None


def get_border_style_from_cell(wb, cell):
    h_styles = {}
    for b_dir in ["right", "left", "top", "bottom"]:
        b_s = getattr(cell.border, b_dir)
        if not b_s:
            continue
        border_style = BORDER_STYLES.get(b_s.style)
        #if border_style is None and b_s.style:
        #    border_style = DEFAULT_BORDER_STYLE

        if not border_style:
            continue

        for k, v in border_style.items():
            h_styles["border-%s-%s" % (b_dir, k)] = v
        if b_s.color:
            h_styles["border-%s-color" % (b_dir)] = normalize_color(wb, b_s.color)
    return h_styles


def get_styles_from_cell(wb, cell, merged_cell_map=None, default_cell_border="none"):
    merged_cell_map = merged_cell_map or {}

    h_styles = {"border-collapse": "collapse"}
    b_styles = get_border_style_from_cell(wb, cell)
    if merged_cell_map:
        # TODO edged_cells
        for m_cell in merged_cell_map["cells"]:
            b_styles.update(get_border_style_from_cell(wb, m_cell))

    #for b_dir in ["border-right", "border-left", "border-top", "border-bottom"]:
    #    style_tag = b_dir + "-style"
    #    if (b_dir not in b_styles) and (style_tag not in b_styles):
    #        b_styles[b_dir] = default_cell_border
        
    h_styles.update(b_styles)

    if cell.alignment.horizontal:
        h_styles["text-align"] = cell.alignment.horizontal
    if cell.alignment.vertical:
        h_styles['vertical-align'] = cell.alignment.vertical

    with contextlib.suppress(AttributeError):
        if cell.fill.patternType == "solid":
            h_styles["background-color"] = normalize_color(wb, cell.fill.fgColor)
        if cell.fill.patternType != "solid" and cell.fill.patternType != None:
            # these are mostly gray
            h_styles["background-color"] = '#EEEEEE'

    if cell.font:
        h_styles["font-size"] = "%spx" % cell.font.sz
        if cell.font.name:
            h_styles["font-family"] = cell.font.name
        if cell.font.color:
            h_styles["color"] = normalize_color(wb, cell.font.color)
        if cell.font.b:
            h_styles["font-weight"] = "bold"
        if cell.font.i:
            h_styles["font-style"] = "italic"
        if cell.font.u:
            h_styles["text-decoration"] = "underline"
    return h_styles


def get_cell_id(cell):
    return "{}".format(cell.coordinate)


def image_to_data(image: Image) -> dict:
    _from: AnchorMarker = image.anchor._from
    graphicalProperties: GraphicalProperties = image.anchor.pic.graphicalProperties
    transform = graphicalProperties.transform
    # http://officeopenxml.com/drwSp-location.php
    offsetX = units.EMU_to_pixels(_from.colOff)
    offsetY = units.EMU_to_pixels(_from.rowOff)
    # TODO recalculate to relative cell
    #import pdb; pdb.set_trace()

    data = {
        "col": _from.col + 1,
        "row": _from.row + 1,
        "offset": {"x": offsetX, "y": offsetY},
        "width": units.EMU_to_pixels(transform.ext.width) if transform else image.width,
        "height": units.EMU_to_pixels(transform.ext.height) if transform else image.height,
        "src": bytes_to_datauri(image.ref, image.path),
        "style": {
            "margin-left": f"{offsetX}px",
            "margin-top": f"{offsetY}px",
        },
    }
    return data


def images_to_data(ws: Worksheet):
    images: List[Image] = ws._images

    images_data = defaultdict(list)
    for _i in images:
        _id = image_to_data(_i)
        images_data[str(_id["col"]) + ":" + str(_id["row"])].append(_id)
    return images_data


def worksheet_to_data(wb, ws, locale=None, fs=None, default_cell_border="none"):
    ws_id = wb.worksheets.index(ws) + 1
    merged_cell_map = {}
    if OPENPYXL_24:
        merged_cell_ranges = ws.merged_cell_ranges
        excluded_cells = set(ws.merged_cells)
    else:
        merged_cell_ranges = [cell_range.coord for cell_range in ws.merged_cells.ranges]
        excluded_cells = set(
            [
                cell
                for cell_range in merged_cell_ranges
                for rows in rows_from_range(cell_range)
                for cell in rows
            ]
        )

    for cell_range in merged_cell_ranges:
        cell_range_list = list(ws[cell_range])
        m_cell = cell_range_list[0][0]

        colspan = len(cell_range_list[0])
        rowspan = len(cell_range_list)
        merged_cell_map[m_cell.coordinate] = {
            "attrs": {
                "colspan": None if colspan <= 1 else colspan,
                "rowspan": None if rowspan <= 1 else rowspan,
            },
            "cells": [c for rows in cell_range_list for c in rows],
        }

        excluded_cells.remove(m_cell.coordinate)

    max_col_number = 0

    data_list = []
    cell_styles = {}
    cell_classnames = {}
    row_height = {}
    no_found_data = True
    cells_with_list = set()

    # keep track of data validation cells
    for dv in ws.data_validations.dataValidation:
      if dv.type == 'list':
         for cellrange in dv.sqref.ranges:
            for row in cellrange.rows:
                for cell_coord in row:
                    cells_with_list.add(cell_coord)

    for row_i, row in  enumerate(reversed(list(ws.iter_rows()))):
        data_row = []
        data_list.insert(0, data_row)
        if no_found_data and all(cell.value == None for cell in row):
            continue
        no_found_data = False

        for col_i, cell in enumerate(row):
            row_dim = ws.row_dimensions[cell.row]
        
            if cell.coordinate in excluded_cells or row_dim.hidden:
                continue
        
            if col_i > max_col_number:
                max_col_number = col_i
        
            height = 19

            if row_dim.customHeight:
                height = round(row_dim.height, 2)
            row_height[cell.row] = height


            f_cell = None
            if fs:
                f_cell = fs[cell.coordinate]
            formatted_value = format_cell(cell, locale=locale, f_cell=f_cell)
            formatted_value = formatted_value.replace("\n", "<br/>") if type(formatted_value) == str else formatted_value    
            cell_data = {
                "column": cell.column,
                "row": cell.row,
                "value": cell.value,
                "formatted_value": formatted_value,
                "is_list": (cell.row, cell.column) in cells_with_list,
                "attrs": {"id": get_cell_id(cell)},
                "style": {},
            }
            merged_cell_info = merged_cell_map.get(cell.coordinate, {})
            if merged_cell_info:
                cell_data["attrs"].update(merged_cell_info["attrs"])
            cell_data["style"].update(
                get_styles_from_cell(wb, cell, merged_cell_info, default_cell_border)
            )
            style_key = json.dumps(cell_data["style"])
            cell_styles[style_key] = cell_styles[style_key] + 1 if style_key in cell_styles else 1
            cell_classnames[style_key] = 'wsid-%s-cellst-%s-%s' % (ws_id, row_i, col_i)
            data_row.append(cell_data)
    col_list = []
    max_col_number += 1

    column_dimensions = sorted(
        ws.column_dimensions.items(), key=lambda d: column_index_from_string(d[0])
    )
    for col_i, col_dim in column_dimensions:
        if not all([col_dim.min, col_dim.max]):
            continue
        
        width = 0.89
        if col_dim.customWidth:
            width = round(col_dim.width / 10.0, 2)
        col_width = 96 * width
        
        for _ in six.moves.range((col_dim.max - col_dim.min) + 1):
            visibility = "collapse" if col_dim.width == 0 else "visible"
            max_col_number -= 1
            col_list.append(
                {
                    "index": col_dim.index,
                    "hidden": col_dim.hidden,
                    "width": col_width,
                    "style": {"min-width": "{}px".format(col_width), "visibility": visibility},
                }
            )
            if max_col_number < 0:
                break
    style_data = {}
    for col in data_list:
        for cell in col:
            style_key = json.dumps(cell['style'])
            if cell_styles[style_key] > 1:
                classname = cell_classnames[style_key]
                style_data[classname] = cell['style']
                cell['attrs'].update({'class': classname})
                cell['style'] = {}

    return {"title": ws.title, "rows": data_list, "cols": col_list, "images": images_to_data(ws), "styles": style_data, "rowHeights": row_height}


def render_table(data, append_headers, append_lineno):
    html = [
        "<table  "
        'style="border-collapse: collapse" '
        'border="0" '
        'cellspacing="0" '
        'cellpadding="0">'
        "<colgroup>"
    ]
    hidden_columns = set()
    for col in data["cols"]:
        if col["hidden"]:
            hidden_columns.add(col["index"])
        html.append(
            '<col {attrs} style="{styles}">'.format(
                attrs=render_attrs(col.get("attrs")),
                styles=render_inline_styles(col.get("style")),
            )
        )
    html.append("</colgroup>")

    append_headers(data, html)

    for i, row in enumerate(data["rows"]):
        trow = ["<tr>"]
        append_lineno(trow, i)
        for cell in row:
            if cell["column"] in hidden_columns:
                continue
            images = []
            colspan = cell["attrs"]["colspan"] if 'colspan' in cell["attrs"] else 1
            colspan = 1 if colspan is None else colspan
            for colspanned_col in range(cell["column"], cell["column"] + colspan):
                images = images + (data["images"].get(str(colspanned_col)+ ':' +str(cell["row"])) or [])
            
 
            formatted_images = []
            for img in images:
                styles = render_inline_styles(img["style"])
                img_tag = (
                    '<img width="{width}" height="{height}"'
                    'style="{styles_str}"'
                    'src="{src}"'
                    "/>"
                ).format(styles_str=styles, **img)
                formatted_images.append(img_tag)
            trow.append(
                (
                    '<td {attrs_str} style="{styles_str}">'
                    "{formatted_value}"
                    "{formatted_images}"
                    "</td>"
                ).format(
                    attrs_str=render_attrs(cell["attrs"]),
                    styles_str=render_inline_styles(cell["style"]),
                    formatted_images="\n".join(formatted_images),
                    **cell,
                )
            )

        trow.append("</tr>")
        html.append("\n".join(trow))
    html.append("</table>")
    return "\n".join(html)

def render_styles(data):
    styles = data['styles']
    style_data = ""
    for key in styles.keys():
        value = styles[key]
        style_data = "%s .%s { %s }" % (style_data, key, render_inline_styles(value))
    return style_data
    

def render_data_to_html(data, append_headers, append_lineno):
    html = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <title>Title</title>
        <style>%s</style>
    </head>
    <body>
        %s
    </body>
    </html>
    """
    return html % (render_styles(data), render_table(data, append_headers, append_lineno))


def get_sheet(wb, sheet):
    ws = wb.active
    if sheet is not None:
        try:
            ws = wb.get_sheet_by_name(sheet)
        except KeyError:
            ws = wb.worksheets[sheet]
    return ws


def xlsx2html(
    filepath,
    output=None,
    locale="en",
    sheet=None,
    parse_formula=False,
    append_headers=(lambda dumb1, dumb2: True),
    append_lineno=(lambda dumb1, dumb2: True),
    default_cell_border="none",
):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = get_sheet(wb, sheet)

    fs = None
    if parse_formula:
        fb = openpyxl.load_workbook(filepath, data_only=False)
        fs = get_sheet(fb, sheet)

    data = worksheet_to_data(
       wb, ws, locale=locale, fs=fs, default_cell_border=default_cell_border
    )
    html = render_data_to_html(data, append_headers, append_lineno)

    if not output:
        output = io.StringIO()
    if isinstance(output, str):
        output = open(output, "w")
    output.write(html)
    output.flush()
    return output
